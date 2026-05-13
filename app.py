import streamlit as st
import pandas as pd
import os
import shutil
import re
import base64
import streamlit.components.v1 as components
from openpyxl import load_workbook
from PIL import Image
import io

# --- 1. Page Configuration ---
st.set_page_config(page_title="Auslite Spec - Designer Edition", layout="wide")

# --- 2. Advanced UI Design Styling ---
st.markdown("""
    <style>
    .main { background-color: #ffffff; }
    .block-container { padding: 2rem 3rem; }
    
    [data-testid="column"]:nth-of-type(1) {
        background-color: #f8f9fa;
        padding: 25px 30px 25px 25px !important;
        border-radius: 12px 0 0 12px;
        border: 1px solid #eeeeee;
        border-right: 1px solid #dddddd !important;
        box-shadow: inset 0 1px 3px rgba(0,0,0,0.02);
    }
    
    [data-testid="column"]:nth-of-type(2) {
        background-color: #ffffff;
        padding-left: 60px !important;
        padding-top: 10px !important;
    }
    
    div.stButton > button { 
        padding: 4px 5px !important; 
        font-size: 10px !important; 
        height: auto !important;
        border: 1px solid #e0e0e0 !important;
        background-color: white !important;
        color: #666 !important;
        transition: all 0.2s ease;
    }
    div.stButton > button:hover {
        border-color: #ff0000 !important;
        color: #ff0000 !important;
        background-color: #fffafa !important;
    }

    .price-card { 
        background-color: #ffffff; 
        border-left: 5px solid #ff0000; 
        padding: 18px 22px; 
        margin-bottom: 20px; 
        box-shadow: 0 4px 12px rgba(0,0,0,0.05);
        border-radius: 4px;
    }
    .price-amount { color: #FF0000; font-size: 38px; font-weight: 900; letter-spacing: -1px; }

    .spec-table { width: 100%; border-collapse: collapse; margin-top: 10px; }
    .spec-table thead th { 
        background-color: #1a1a1a; color: white; text-align: left; 
        padding: 12px 15px; font-size: 11px; text-transform: uppercase; letter-spacing: 1px;
    }
    .spec-table tbody td { padding: 14px 15px; color: #333; border-bottom: 1px solid #f0f0f0; font-size: 13px; }

    .gallery-container {
        width: 100%;
        aspect-ratio: 1 / 1; 
        background-color: #ffffff; 
        border-radius: 8px;
        display: flex;
        align-items: center;
        justify-content: center;
        overflow: hidden;
        border: 1px solid #eeeeee;
        margin-bottom: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.03);
    }
    .gallery-container img {
        max-width: 88%;
        max-height: 88%;
        object-fit: contain; 
    }
    </style>
    """, unsafe_allow_html=True)

# --- 🚀 核心辅助函数 ---
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(s))]

def get_image_base64(path):
    if os.path.exists(path):
        try:
            with open(path, "rb") as f: return base64.b64encode(f.read()).decode('utf-8')
        except: return ""
    return ""

# --- 3. Robust Data Processing ---
ORIGINAL_FILE = "Test.xlsx"
IMAGE_DIR = "product_images"

@st.cache_resource
def extract_images_robustly():
    """扫描所有工作表提取图片"""
    try:
        if not os.path.exists(IMAGE_DIR): os.makedirs(IMAGE_DIR, exist_ok=True) 
        if not os.path.exists(ORIGINAL_FILE): return
        
        pxl_doc = load_workbook(ORIGINAL_FILE)
        
        for sheet in pxl_doc.worksheets:
            # 找到当前 Sheet 的 CODE 列索引
            code_col_idx = None
            for col in range(1, sheet.max_column + 1):
                cell_val = str(sheet.cell(row=1, column=col).value).upper()
                if "CODE" in cell_val:
                    code_col_idx = col
                    break
            
            if not code_col_idx: continue

            # 建立行号与 CODE 的映射
            code_map = {}
            for row in range(2, sheet.max_row + 1):
                val = sheet.cell(row=row, column=code_col_idx).value 
                if val: code_map[row] = str(val).strip()

            if hasattr(sheet, '_images'):
                for img in sheet._images:
                    try:
                        row_idx = img.anchor._from.row + 1 
                        code_name = code_map.get(row_idx)
                        if code_name:
                            pil_img = None
                            if hasattr(img, '_data'): 
                                pil_img = Image.open(io.BytesIO(img._data()))
                            elif hasattr(img, 'image'): 
                                pil_img = img.image
                            
                            if pil_img:
                                if pil_img.mode in ("RGBA", "P"): pil_img = pil_img.convert("RGB")
                                # 强制保存为 PNG 以便后续读取
                                pil_img.save(f"{IMAGE_DIR}/{code_name}.png", "PNG")
                    except: continue
        pxl_doc.close()
    except Exception as e:
        st.error(f"Image Extraction Error: {e}")

@st.cache_data
def get_all_categories():
    """获取所有 Sheet 名字"""
    xl = pd.ExcelFile(ORIGINAL_FILE)
    return xl.sheet_names

@st.cache_data
def load_category_data(sheet_name):
    """加载特定品类，去掉了 DALI 过滤"""
    df = pd.read_excel(ORIGINAL_FILE, sheet_name=sheet_name)
    df.columns = [str(c).upper().strip().replace('\n', '') for c in df.columns]
    if "MODEL" in df.columns: 
        df["MODEL"] = df["MODEL"].ffill()
    return df

# --- 4. Main Application ---
try:
    # 提取所有图片
    extract_images_robustly()
    # 获取分类
    all_sheets = get_all_categories()

    # --- 布局比例 ---
    left_col, right_col = st.columns([3, 5]) 

    # --- 【左侧：分类切换 + 系列画廊】 ---
    with left_col:
        st.markdown("<h3 style='font-weight:800; margin-bottom:10px;'>Gallery</h3>", unsafe_allow_html=True)
        
        # 分类切换按钮 (红圈位置)
        selected_cat = st.pills("Select Category:", all_sheets, selection_mode="single", default=all_sheets[0])
        
        if not selected_cat:
            selected_cat = all_sheets[0]

        # 加载选中分类数据
        df = load_category_data(selected_cat)
        unique_series = sorted(df["MODEL"].dropna().unique(), key=natural_sort_key)
        
        # 独立存储每个分类选中的系列
        state_key = f"active_{selected_cat}"
        if state_key not in st.session_state:
            st.session_state[state_key] = unique_series[0] if unique_series else None

        st.markdown("---")
        
        # 渲染九宫格
        grid_cols = 4 
        for i in range(0, len(unique_series), grid_cols):
            cols = st.columns(grid_cols)
            for j in range(grid_cols):
                idx = i + j
                if idx < len(unique_series):
                    s_name = unique_series[idx]
                    # 获取该系列第一行的 CODE 来展示缩略图
                    s_data = df[df["MODEL"] == s_name]
                    c_code = str(s_data.iloc[0]["CODE"]).strip()
                    path = f"{IMAGE_DIR}/{c_code}.png"
                    
                    with cols[j]:
                        img_b64 = get_image_base64(path)
                        if img_b64:
                            st.markdown(f'<div class="gallery-container"><img src="data:image/png;base64,{img_b64}"></div>', unsafe_allow_html=True)
                        else:
                            st.markdown('<div class="gallery-container" style="font-size:9px; color:#999;">No Img</div>', unsafe_allow_html=True)
                        
                        is_act = (st.session_state[state_key] == s_name)
                        if st.button(s_name, key=f"btn_{selected_cat}_{s_name}", use_container_width=True, type="primary" if is_act else "secondary"):
                            st.session_state[state_key] = s_name
                            st.rerun()

    # --- 【右侧：详情显示区】 ---
    with right_col:
        active_series = st.session_state.get(state_key)
        
        if active_series:
            st.markdown(f"<h2 style='font-weight:900; color:#1a1a1a; margin-top:0;'>📦 {active_series}</h2>", unsafe_allow_html=True)
            
            s_df = df[df["MODEL"] == active_series]
            c_list = sorted(s_df["CODE"].dropna().unique(), key=natural_sort_key)
            sel_c = st.selectbox("Variant Select:", c_list)
            res_r = s_df[s_df["CODE"] == sel_c].iloc[0]
            
            c1, c2 = st.columns([0.8, 2])
            det_path = f"{IMAGE_DIR}/{str(sel_c).strip()}.png"
            i_b64 = get_image_base64(det_path)

            with c1:
                if os.path.exists(det_path):
                    st.image(det_path, use_container_width=True)
                else:
                    st.warning("No Variant Image.")
            
            with c2:
                # 价格显示
                price_col = next((c for c in df.columns if "PRICE" in c), None)
                if price_col:
                    p_v = str(res_r[price_col]).strip()
                    st.markdown(f'<div class="price-card"><div style="font-size:12px; font-weight:bold; color:#666; letter-spacing:1px;">PRICE UNIT</div><div class="price-amount">{p_v if p_v.startswith("$") else "$"+p_v}</div></div>', unsafe_allow_html=True)

                # 复制按钮逻辑
                order = ['MODEL', 'CODE', 'COLOR', 'WATTAGE', 'CCT', 'CRI', 'DEGREE', 'DIM METHOD', 'IP', 'DIMENSION', 'CUTOUT', 'INSTALLATION', 'PRICE']
                copy_v = []
                for k in order:
                    m = next((c for c in res_r.index if c.upper() == k.upper()), None)
                    copy_v.append(str(res_r[m]) if m else "")
                
                components.html(f"""
                    <script>
                    function copyT() {{ navigator.clipboard.writeText(`{"\\t".join(copy_v)}`); alert('✅ Data Copied!'); }}
                    async function copyI() {{
                        const b = "{i_b64}"; if(!b) return;
                        const blob = await (await fetch('data:image/png;base64,'+b)).blob();
                        await navigator.clipboard.write([new ClipboardItem({{'image/png': blob}})]);
                        alert('🖼️ Image Copied!');
                    }}
                    </script>
                    <div style="display:flex; gap:10px;">
                        <button onclick="copyT()" style="background:#007bff;color:white;border:none;padding:10px 18px;border-radius:6px;cursor:pointer;font-weight:bold;font-size:13px;">📋 Copy Row Data</button>
                        <button onclick="copyI()" style="background:#28a745;color:white;border:none;padding:10px 18px;border-radius:6px;cursor:pointer;font-weight:bold;font-size:13px;">🖼️ Copy Image</button>
                    </div>
                """, height=60)

            st.markdown("<br>", unsafe_allow_html=True)
            # 规格参数表
            d_list = []
            for o in order:
                m = next((c for c in res_r.index if c.upper() == o.upper()), None)
                d_list.append({"Key": o, "Value": res_r[m] if m else "-"})
            
            h_t = '<div style="overflow-x:auto;"><table class="spec-table"><thead><tr>'
            for col in d_list: h_t += f'<th>{col["Key"]}</th>'
            h_t += '</tr></thead><tbody><tr>'
            for col in d_list: h_t += f'<td>{col["Value"]}</td>'
            h_t += '</tr></tbody></table></div>'
            st.markdown(h_t, unsafe_allow_html=True)
        else:
            st.info("Select a category and series to view details.")

except Exception as e:
    st.error(f"App Error: {e}")